import re
from openpyxl import load_workbook, Workbook


# train.xlsx 를 읽어서 Mr, Miss, Mrs, Others 분류한 후 시트로
# 저장

wb = load_workbook("./crawl/regex/train.xlsx")
ws1 = wb.active

# 새로운 엑셀 파일
new_wb = Workbook()

# Mr
ws1_man = new_wb.active
ws1_man.title = "남성"
ws1_man.column_dimensions["D"].width = 70

# Miss
ws2_miss = new_wb.create_sheet()
ws2_miss.title = "미혼여성"
ws2_miss.column_dimensions["D"].width = 70

# Mrs
ws3_mrs = new_wb.create_sheet()
ws3_mrs.title = "기혼여성"
ws3_mrs.column_dimensions["D"].width = 70

# Others
ws4_others = new_wb.create_sheet()
ws4_others.title = "기타"
ws4_others.column_dimensions["D"].width = 70


# 공백Mr.  Miss.  Mrs.

pattern = re.compile(" [A-Za-z]+\.")

for each_row in ws1.rows:
    # name 가져온 후 패턴과 일치하는지 확인
    data = pattern.findall(each_row[3].value)  # [' Mr.']

    # 원본 엑셀에 있는 타이틀 복사 후 새 엑셀파일로 붙여넣기

    # list1 = []
    # for title in row:
    #     list1.append(title.value)

    # ws1_man.append(list1)

    if each_row[0].row == 1:
        ws1_man.append([each_item.value for each_item in each_row])
        ws2_miss.append([each_item.value for each_item in each_row])
        ws3_mrs.append([each_item.value for each_item in each_row])
        ws4_others.append([each_item.value for each_item in each_row])
    else:
        if len(data) > 0:
            if data[0] == " Mr.":
                ws1_man.append([each_item.value for each_item in each_row])
            elif data[0] == " Miss.":
                ws2_miss.append([each_item.value for each_item in each_row])
            elif data[0] == " Mrs.":
                ws3_mrs.append([each_item.value for each_item in each_row])
            else:
                ws4_others.append([each_item.value for each_item in each_row])

new_wb.save("./crawl/regex/train_gender.xlsx")

wb.close()
new_wb.close()
